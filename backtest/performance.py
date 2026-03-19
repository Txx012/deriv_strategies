# -*- coding: utf-8 -*-
"""绩效分析器：所有策略通用，计算夏普/最大回撤等指标+绘制专业图表+导出Excel报告"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict, Optional
from utils.log_utils import logger

# 全局绘图配置（中文字体+样式）
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.figsize'] = (16, 12)
plt.rcParams['axes.grid'] = True
plt.rcParams['grid.linestyle'] = '--'
plt.rcParams['grid.alpha'] = 0.7

class PerformanceAnalyzer:
    def __init__(self, result_df: pd.DataFrame, backtest_config: Dict, strategy_name: str):
        self.result_df = result_df
        self.cfg = backtest_config
        self.strategy_name = strategy_name
        self.trading_days = len(result_df)
        self.annual_trading_days = self.cfg.get("annual_trading_days", 252)
        self.risk_free_rate = self.cfg.get("risk_free_rate", 0.02)  # 无风险利率2%
        self.metrics: Dict = {}
        self.logger = logger
        self._preprocess_data()  # 初始化自动预处理

    def _preprocess_data(self):
        """数据预处理：计算日收益率/累计收益/回撤曲线等基础指标"""
        if self.result_df.empty:
            raise ValueError("回测结果为空，无法进行绩效分析")
        # 时间格式标准化
        self.result_df["datetime"] = pd.to_datetime(self.result_df["datetime"])
        self.result_df = self.result_df.sort_values("datetime").reset_index(drop=True)
        # 计算日收益率（空值填0）
        self.result_df["daily_return"] = self.result_df["capital"].pct_change().fillna(0)
        # 累计收益曲线（复利）
        self.result_df["cum_return_curve"] = (1 + self.result_df["daily_return"]).cumprod() - 1
        # 滚动20日波动率（年化）
        self.result_df["rolling_20d_vol"] = self.result_df["daily_return"].rolling(20).std() * np.sqrt(self.annual_trading_days)
        # 历史最大净资产&最大回撤曲线
        self.result_df["max_capital"] = self.result_df["capital"].cummax()
        self.result_df["drawdown_curve"] = (self.result_df["capital"] - self.result_df["max_capital"]) / self.result_df["max_capital"]
        # 区分盈利/亏损日
        self.result_df["is_win"] = self.result_df["daily_return"] > 0
        logger.info(f"绩效数据预处理完成，有效交易天数：{self.trading_days}")

    def calculate_metrics(self) -> Dict:
        """计算全量核心绩效指标（4大类18项，覆盖收益/风险/风险收益比/交易质量）"""
        dr = self.result_df["daily_return"]
        cr = self.result_df["cum_return_curve"].iloc[-1]
        dd = self.result_df["drawdown_curve"]
        final_cap = self.result_df["capital"].iloc[-1]
        initial_cap = self.cfg["initial_capital"]

        # 1. 基础收益指标
        total_pnl = final_cap - initial_cap
        annual_return = cr * (self.annual_trading_days / self.trading_days)
        daily_return_mean = dr.mean()
        monthly_return = annual_return / 12

        # 2. 风险指标
        annual_vol = dr.std() * np.sqrt(self.annual_trading_days)
        max_drawdown = abs(dd.min())
        drawdown_max_days = self._calc_max_drawdown_days()  # 最大回撤持续天数
        var_95 = np.percentile(dr, 5)  # 95%置信区间VaR
        cvar_95 = dr[dr <= var_95].mean()  # 条件风险价值

        # 3. 风险收益比指标（核心）
        sharpe = (annual_return - self.risk_free_rate) / annual_vol if annual_vol != 0 else 0
        sortino = (annual_return - self.risk_free_rate) / (dr[dr < 0].std() * np.sqrt(self.annual_trading_days)) if dr[dr < 0].std() != 0 else 0
        calmar = annual_return / max_drawdown if max_drawdown != 0 else np.inf
        omega = (dr[dr > 0].sum() - self.risk_free_rate/self.annual_trading_days * self.trading_days) / abs(dr[dr < 0].sum()) if dr[dr < 0].sum() != 0 else np.inf

        # 4. 交易质量指标
        win_rate = self.result_df["is_win"].mean()
        profit_factor = abs(dr[dr > 0].sum() / dr[dr < 0].sum()) if dr[dr < 0].sum() != 0 else np.inf
        avg_win = dr[dr > 0].mean()
        avg_loss = abs(dr[dr < 0].mean())
        win_loss_ratio = avg_win / avg_loss if avg_loss != 0 else np.inf

        # 封装所有指标（保留4位小数，百分比转小数）
        self.metrics = {
            # 基础信息
            "策略名称": self.strategy_name,
            "初始资金(元)": initial_cap,
            "最终资金(元)": round(final_cap, 2),
            "交易天数": self.trading_days,
            # 收益类
            "总收益(元)": round(total_pnl, 2),
            "累计收益率": round(cr, 4),
            "年化收益率": round(annual_return, 4),
            "月化收益率": round(monthly_return, 4),
            "日均收益率": round(daily_return_mean, 6),
            # 风险类
            "年化波动率": round(annual_vol, 4),
            "最大回撤": round(max_drawdown, 4),
            "最大回撤持续天数": drawdown_max_days,
            "95%VaR": round(var_95, 6),
            "95%CVaR": round(cvar_95, 6),
            # 风险收益比类
            "夏普比率": round(sharpe, 2),
            "索提诺比率": round(sortino, 2),
            "卡玛比率": round(calmar, 2),
            "欧米茄比率": round(omega, 2),
            # 交易质量类
            "胜率": round(win_rate, 4),
            "盈利因子": round(profit_factor, 2),
            "平均盈利/平均亏损": round(win_loss_ratio, 2),
            "平均单日盈利": round(avg_win, 6),
            "平均单日亏损": round(avg_loss, 6)
        }
        self._print_metrics()  # 打印指标到控制台
        return self.metrics

    def _calc_max_drawdown_days(self) -> int:
        """计算最大回撤的持续天数（辅助方法）"""
        drawdown = self.result_df["drawdown_curve"]
        drawdown_periods = []
        start = None
        for idx, val in enumerate(drawdown):
            if val < 0 and start is None:
                start = idx
            elif val == 0 and start is not None:
                drawdown_periods.append(idx - start)
                start = None
        # 处理未结束的回撤
        if start is not None:
            drawdown_periods.append(len(drawdown) - start)
        return max(drawdown_periods) if drawdown_periods else 0

    def _print_metrics(self):
        """格式化打印绩效指标（控制台友好展示）"""
        logger.info("\n" + "="*50)
        logger.info(f"          {self.strategy_name} - 回测绩效指标汇总")
        logger.info("="*50)
        for k, v in self.metrics.items():
            # 百分比字段特殊格式化
            if k in ["累计收益率", "年化收益率", "月化收益率", "胜率", "最大回撤", "年化波动率"]:
                logger.info(f"{k:<15}: {v:.2%}")
            elif "元" in k or k in ["交易天数", "最大回撤持续天数"]:
                logger.info(f"{k:<15}: {v:,}")
            else:
                logger.info(f"{k:<15}: {v:.2f}")
        logger.info("="*50 + "\n")

    def plot_charts(self, save_fig: bool = True):
        """绘制专业回测图表（4个子图：净值+累计收益+回撤+日收益率分布）"""
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle(f"{self.strategy_name} 回测结果可视化", fontsize=20, fontweight="bold", y=0.98)
        c1, c2, c3, c4 = "#2E86AB", "#E63946", "#FCA311", "#1D3557"  # 配色方案

        # 子图1：净资产曲线
        ax1.plot(self.result_df["datetime"], self.result_df["capital"], color=c1, linewidth=2.5, label="净资产")
        ax1.plot(self.result_df["datetime"], self.result_df["max_capital"], color=c2, linewidth=1.5, linestyle="--", label="历史最大净资产")
        ax1.set_title("净资产曲线", fontsize=16, fontweight="bold", pad=10)
        ax1.set_ylabel("资金（元）", fontsize=12)
        ax1.legend(fontsize=10)
        ax1.tick_params(axis="x", rotation=45)

        # 子图2：累计收益率曲线
        ax2.plot(self.result_df["datetime"], self.result_df["cum_return_curve"], color=c1, linewidth=2.5, label="累计收益率")
        ax2.axhline(y=0, color="black", linewidth=1, linestyle="-", alpha=0.5)
        ax2.set_title("累计收益率曲线", fontsize=16, fontweight="bold", pad=10)
        ax2.set_ylabel("收益率", fontsize=12)
        ax2.legend(fontsize=10)
        ax2.tick_params(axis="x", rotation=45)

        # 子图3：最大回撤曲线
        ax3.fill_between(self.result_df["datetime"], self.result_df["drawdown_curve"], 0, color=c2, alpha=0.3, label="回撤")
        ax3.plot(self.result_df["datetime"], self.result_df["drawdown_curve"], color=c2, linewidth=2, label="回撤曲线")
        ax3.set_title("最大回撤曲线", fontsize=16, fontweight="bold", pad=10)
        ax3.set_ylabel("回撤率", fontsize=12)
        ax3.legend(fontsize=10)
        ax3.tick_params(axis="x", rotation=45)

        # 子图4：日收益率分布+核密度估计
        ax4.hist(self.result_df["daily_return"], bins=50, color=c3, alpha=0.6, edgecolor="black", label="日收益率分布")
        self.result_df["daily_return"].plot(kind="kde", color=c4, linewidth=2.5, ax=ax4, label="核密度曲线")
        ax4.axvline(x=0, color="black", linewidth=1, linestyle="-", alpha=0.5)
        ax4.set_title("日收益率分布", fontsize=16, fontweight="bold", pad=10)
        ax4.set_xlabel("日收益率", fontsize=12)
        ax4.set_ylabel("频次", fontsize=12)
        ax4.legend(fontsize=10)

        # 调整布局
        plt.tight_layout(rect=[0, 0, 1, 0.95])
        # 保存图片
        if save_fig:
            fig_dir = os.path.join(self.cfg["result_path"], "charts")
            os.makedirs(fig_dir, exist_ok=True)
            fig_path = os.path.join(fig_dir, f"{self.strategy_name}_backtest_chart.png")
            plt.savefig(fig_path, dpi=300, bbox_inches="tight")
            logger.info(f"回测图表已保存至：{fig_path}")
        plt.show()

    def export_report(self, report_name: Optional[str] = None) -> str:
        """导出详细回测报告（Excel格式，含指标+原始数据+交易记录）"""
        if not self.metrics:
            self.calculate_metrics()  # 未计算指标则先计算
        report_name = report_name or f"{self.strategy_name}_回测报告"
        report_path = os.path.join(self.cfg["result_path"], f"{report_name}.xlsx")
        # 创建Excel工作簿
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            # Sheet1：绩效指标汇总
            metrics_df = pd.DataFrame(list(self.metrics.items()), columns=["指标名称", "指标值"])
            metrics_df.to_excel(writer, sheet_name="绩效指标", index=False)
            # Sheet2：原始回测数据
            self.result_df.to_excel(writer, sheet_name="原始数据", index=False)
            # Sheet3：日收益率详情
            return_df = self.result_df[["datetime", "capital", "daily_return", "cum_return_curve", "drawdown_curve"]]
            return_df.to_excel(writer, sheet_name="收益率详情", index=False)
        # 美化Excel（可选，提升可读性）
        self._beautify_excel(report_path)
        logger.info(f"详细回测报告已导出至：{report_path}")
        return report_path

    def _beautify_excel(self, file_path: str):
        """美化Excel报告（设置字体/对齐/底色，辅助方法）"""
        try:
            wb = Workbook()
            wb = pd.ExcelFile(file_path).book
            # 样式配置
            header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
            cell_alignment = Alignment(horizontal="center", vertical="center")
            # 遍历所有sheet美化
            for ws in wb.worksheets:
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = cell_alignment
                # 自动调整列宽
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2
            wb.save(file_path)
        except Exception as e:
            logger.warning(f"Excel报告美化失败：{e}，不影响数据有效性")

# 测试代码
if __name__ == "__main__":
    import yaml
    import os
    cfg_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config", "backtest_config.yaml")
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    # 生成测试数据
    test_df = pd.DataFrame({
        "datetime": pd.date_range(start="2023-01-01", periods=100, freq="D"),
        "capital": np.cumsum(np.random.randn(100)*100 + 500) + 1000000
    })
    # 初始化分析器
    analyzer = PerformanceAnalyzer(test_df, cfg, "测试策略")
    analyzer.calculate_metrics()
    analyzer.plot_charts()
    analyzer.export_report()